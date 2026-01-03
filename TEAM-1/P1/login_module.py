"""
登录验证模块
负责处理用户登录验证逻辑
"""

import json
import os
from datetime import datetime
from openpyxl import load_workbook
from algorithm import CalculateMd5

# 配置常量
SHARE_FILE_PATH = 'share.txt'
DATA_FILE_PATH = 'DATA.xlsx'


def ReadShareFile():
    """
    从共享文件读取登录数据
    传入值: 无
    返回值: dict - 包含用户名、密码等信息的字典，失败返回None
    """
    try:
        if not os.path.exists(SHARE_FILE_PATH):
            print(f"共享文件不存在: {SHARE_FILE_PATH}")
            return None
        
        with open(SHARE_FILE_PATH, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        return data
    
    except Exception as e:
        print(f"读取共享文件错误: {e}")
        return None


def ValidateUserCredentials(pre_user_name, pre_user_psw):
    """
    验证用户凭据
    传入值: pre_user_name (str) - 用户名
            pre_user_psw (str) - 密码
    返回值: tuple - (bool, str, str) 验证结果、用户IP、错误消息
    """
    try:
        # 计算密码的MD5哈希值
        psw_md5 = CalculateMd5(pre_user_psw)
        
        if not psw_md5:
            return False, None, "密码加密失败"
        
        print(f"密码MD5值: {psw_md5}")
        
        # 读取DATA.xlsx文件
        if not os.path.exists(DATA_FILE_PATH):
            return False, None, f"数据文件不存在: {DATA_FILE_PATH}"
        
        workbook = load_workbook(DATA_FILE_PATH)
        sheet = workbook.active
        
        # 遍历Excel查找匹配的用户
        # 假设Excel格式：第一行是标题，第一列是用户名，第二列是MD5密码，第三列是IP，第四列是最后登录时间
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
            stored_username = row[0].value
            stored_password_md5 = row[1].value
            stored_ip = row[2].value if len(row) > 2 and row[2].value else "未知"
            
            # 比较用户名和MD5密码
            if stored_username == pre_user_name and stored_password_md5 == psw_md5:
                print(f"用户验证成功: {pre_user_name}")
                return True, stored_ip, row_idx
        
        workbook.close()
        return False, None, "用户名或密码错误"
    
    except Exception as e:
        print(f"验证用户凭据错误: {e}")
        return False, None, f"验证过程出错: {str(e)}"


def UpdateLoginInfo(row_idx, login_ip):
    """
    更新用户的登录信息
    传入值: row_idx (int) - Excel中用户所在行号
            login_ip (str) - 登录IP地址
    返回值: bool - 更新成功返回True，失败返回False
    """
    try:
        workbook = load_workbook(DATA_FILE_PATH)
        sheet = workbook.active
        
        # 获取当前时间
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # 更新IP（第3列）和最后登录时间（第4列）
        sheet.cell(row=row_idx, column=3, value=login_ip)
        sheet.cell(row=row_idx, column=4, value=current_time)
        
        # 保存文件
        workbook.save(DATA_FILE_PATH)
        workbook.close()
        
        print(f"登录信息已更新: IP={login_ip}, 时间={current_time}")
        return True
    
    except Exception as e:
        print(f"更新登录信息错误: {e}")
        return False


def ProcessLogin():
    """
    处理登录流程的主函数
    传入值: 无
    返回值: dict - 包含登录结果的字典
    """
    result = {
        'ret_status': 'ret_ERR',
        'ret_message': '登录失败',
        'ret_user_name': '',
        'ret_login_time': ''
    }
    
    try:
        print("\n" + "="*60)
        print("开始处理登录验证")
        print("="*60 + "\n")
        
        # 1. 读取共享文件中的用户数据
        print("【步骤1】读取共享文件...")
        login_data = ReadShareFile()
        
        if not login_data:
            result['ret_message'] = "无法读取登录数据"
            return result
        
        pre_user_name = login_data.get('pre_user_name', '')
        pre_user_psw = login_data.get('pre_user_psw', '')
        pre_cookie = login_data.get('pre_cookie', '')
        
        print(f"用户名: {pre_user_name}")
        print(f"Cookie: {pre_cookie}\n")
        
        if not pre_user_name or not pre_user_psw:
            result['ret_message'] = "用户名或密码为空"
            return result
        
        # 2. 调用MD5算法验证密码
        print("【步骤2】验证用户凭据...")
        is_valid, user_ip, row_or_msg = ValidateUserCredentials(pre_user_name, pre_user_psw)
        
        if not is_valid:
            result['ret_message'] = row_or_msg
            print(f"✗ 验证失败: {row_or_msg}\n")
            return result
        
        print("✓ 凭据验证成功\n")
        
        # 3. 更新登录信息
        print("【步骤3】更新登录信息...")
        
        # 尝试从cookie中提取IP（简化处理，实际应该从HTTP请求获取）
        login_ip = "127.0.0.1"  # 默认IP
        if 'ip=' in pre_cookie:
            try:
                login_ip = pre_cookie.split('ip=')[1].split(';')[0]
            except:
                pass
        
        update_success = UpdateLoginInfo(row_or_msg, login_ip)
        
        if update_success:
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            result['ret_status'] = 'ret_OK'
            result['ret_message'] = '登录成功'
            result['ret_user_name'] = pre_user_name
            result['ret_login_time'] = current_time
            print("✓ 登录信息更新成功\n")
        else:
            result['ret_message'] = '登录成功但信息更新失败'
            print("✗ 登录信息更新失败\n")
        
        print("="*60)
        print(f"登录结果: {result['ret_status']}")
        print("="*60 + "\n")
        
        return result
    
    except Exception as e:
        print(f"登录处理错误: {e}")
        result['ret_message'] = f"登录处理异常: {str(e)}"
        return result


def WriteLoginResult(result):
    """
    将登录结果写入文件
    传入值: result (dict) - 登录结果字典
    返回值: bool - 写入成功返回True，失败返回False
    """
    try:
        result_file = 'login_result.txt'
        with open(result_file, 'w', encoding='utf-8') as f:
            json.dump(result, f, ensure_ascii=False, indent=4)
        return True
    
    except Exception as e:
        print(f"写入登录结果错误: {e}")
        return False


def InitializeDataFile():
    """
    初始化DATA.xlsx文件（如果不存在或为空）
    传入值: 无
    返回值: NULL
    """
    try:
        from openpyxl import Workbook
        
        if not os.path.exists(DATA_FILE_PATH):
            print("创建新的DATA.xlsx文件...")
            workbook = Workbook()
            sheet = workbook.active
            
            # 设置标题行
            sheet['A1'] = '用户名'
            sheet['B1'] = '密码(MD5)'
            sheet['C1'] = 'IP地址'
            sheet['D1'] = '最后登录时间'
            
            # 添加测试用户（用户名: admin, 密码: 1234567）
            sheet['A2'] = 'admin'
            sheet['B2'] = CalculateMd5('1234567')  # 密码1234567的MD5
            sheet['C2'] = '127.0.0.1'
            sheet['D2'] = '2026-01-01 00:00:00'
            
            # 添加测试用户（用户名: user123, 密码: password123）
            sheet['A3'] = 'user123'
            sheet['B3'] = CalculateMd5('password123')
            sheet['C3'] = '127.0.0.1'
            sheet['D3'] = '2026-01-01 00:00:00'
            
            workbook.save(DATA_FILE_PATH)
            workbook.close()
            print("DATA.xlsx文件创建成功")
            print("测试账号1: 用户名=admin, 密码=1234567")
            print("测试账号2: 用户名=user123, 密码=password123")
    
    except Exception as e:
        print(f"初始化数据文件错误: {e}")


if __name__ == '__main__':
    # 初始化数据文件（如果需要）
    InitializeDataFile()
    
    # 处理登录
    result = ProcessLogin()
    
    # 写入结果
    WriteLoginResult(result)
    
    # 打印最终结果
    print("\n最终登录结果:")
    print(json.dumps(result, ensure_ascii=False, indent=2))
