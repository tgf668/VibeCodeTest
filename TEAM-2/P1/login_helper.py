"""
login_helper.py - 登录辅助模块
为login.c提供Python功能支持（MD5计算、Excel读写）
"""

import sys
import os
from datetime import datetime

# 导入算法模块
from algorithm import CalculateMd5

# 尝试导入openpyxl，用于Excel操作
try:
    from openpyxl import Workbook, load_workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("警告: openpyxl未安装，Excel功能不可用")

# 常量定义
DATA_FILE_PATH = "DATA.xlsx"


def ProcessMd5Command():
    """
    传入值: NULL
    返回值: NULL
    处理MD5计算命令，从temp_md5_input.txt读取数据，结果写入temp_md5_output.txt
    """
    try:
        # 读取输入
        with open("temp_md5_input.txt", "r", encoding="utf-8") as file:
            pre_data = file.read()
        
        # 计算MD5
        ret_md5_hash = CalculateMd5(pre_data)
        
        # 写入输出
        with open("temp_md5_output.txt", "w", encoding="utf-8") as file:
            file.write(ret_md5_hash)
        
        return 0
    
    except Exception as e:
        print(f"MD5计算错误: {e}")
        return 1


def ProcessVerifyCommand():
    """
    传入值: NULL
    返回值: NULL
    处理验证命令，从temp_verify_input.txt读取用户名和MD5哈希，
    与DATA.xlsx中的数据比较，结果写入temp_verify_output.txt
    """
    try:
        # 读取输入
        with open("temp_verify_input.txt", "r", encoding="utf-8") as file:
            lines = file.readlines()
        
        if len(lines) < 2:
            with open("temp_verify_output.txt", "w", encoding="utf-8") as file:
                file.write("ERROR")
            return 1
        
        pre_user_name = lines[0].strip()
        pre_md5_hash = lines[1].strip()
        
        # 验证用户数据
        is_valid = VerifyUserInExcel(pre_user_name, pre_md5_hash)
        
        # 写入输出
        with open("temp_verify_output.txt", "w", encoding="utf-8") as file:
            file.write("OK" if is_valid else "ERROR")
        
        return 0 if is_valid else 1
    
    except Exception as e:
        print(f"验证错误: {e}")
        with open("temp_verify_output.txt", "w", encoding="utf-8") as file:
            file.write("ERROR")
        return 1


def ProcessUpdateCommand():
    """
    传入值: NULL
    返回值: NULL
    处理更新命令，从temp_update_input.txt读取用户名和IP，
    在DATA.xlsx中更新登录时间和IP
    """
    try:
        # 读取输入
        with open("temp_update_input.txt", "r", encoding="utf-8") as file:
            lines = file.readlines()
        
        if len(lines) < 2:
            return 1
        
        pre_user_name = lines[0].strip()
        pre_ip = lines[1].strip()
        
        # 更新登录记录
        ret_success = UpdateLoginRecordInExcel(pre_user_name, pre_ip)
        
        return 0 if ret_success else 1
    
    except Exception as e:
        print(f"更新错误: {e}")
        return 1


def VerifyUserInExcel(pre_user_name, pre_md5_hash):
    """
    传入值: pre_user_name - 用户名 (str)
            pre_md5_hash - 密码的MD5哈希值 (str)
    返回值: bool - 验证通过返回True，失败返回False
    """
    if not EXCEL_AVAILABLE:
        print("Excel功能不可用")
        return False
    
    if not os.path.exists(DATA_FILE_PATH):
        # 如果文件不存在，创建示例数据文件
        CreateSampleDataFile()
    
    try:
        workbook = load_workbook(DATA_FILE_PATH)
        sheet = workbook.active
        
        # 遍历所有行查找用户
        # 假设格式: 用户名 | 密码MD5 | 最后登录时间 | 最后登录IP
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == pre_user_name and row[1] == pre_md5_hash:
                workbook.close()
                return True
        
        workbook.close()
        return False
    
    except Exception as e:
        print(f"读取Excel错误: {e}")
        return False


def UpdateLoginRecordInExcel(pre_user_name, pre_ip):
    """
    传入值: pre_user_name - 用户名 (str)
            pre_ip - 登录IP地址 (str)
    返回值: bool - 更新成功返回True，失败返回False
    """
    if not EXCEL_AVAILABLE:
        print("Excel功能不可用")
        return False
    
    if not os.path.exists(DATA_FILE_PATH):
        return False
    
    try:
        workbook = load_workbook(DATA_FILE_PATH)
        sheet = workbook.active
        
        # 获取当前时间
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # 查找并更新用户记录
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] == pre_user_name:
                # 更新最后登录时间（第3列）和IP（第4列）
                sheet.cell(row=row_idx, column=3, value=current_time)
                sheet.cell(row=row_idx, column=4, value=pre_ip)
                workbook.save(DATA_FILE_PATH)
                workbook.close()
                return True
        
        workbook.close()
        return False
    
    except Exception as e:
        print(f"更新Excel错误: {e}")
        return False


def CreateSampleDataFile():
    """
    传入值: NULL
    返回值: NULL
    创建示例DATA.xlsx文件
    """
    if not EXCEL_AVAILABLE:
        return
    
    try:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "用户数据"
        
        # 添加表头
        sheet['A1'] = "用户名"
        sheet['B1'] = "密码MD5"
        sheet['C1'] = "最后登录时间"
        sheet['D1'] = "最后登录IP"
        
        # 添加示例用户 (密码为 "pass123" 的MD5 - 7位密码符合要求)
        # MD5("pass123") = 32250170a0dca92d53ec9624f336ca24
        sheet['A2'] = "admin"
        sheet['B2'] = "32250170a0dca92d53ec9624f336ca24"
        sheet['C2'] = ""
        sheet['D2'] = ""
        
        # 添加第二个示例用户 (密码为 "test1234" 的MD5 - 8位密码符合要求)
        # MD5("test1234") = 16d7a4fca7442dda3ad93c9a726597e4
        sheet['A3'] = "test"
        sheet['B3'] = "16d7a4fca7442dda3ad93c9a726597e4"
        sheet['C3'] = ""
        sheet['D3'] = ""
        
        workbook.save(DATA_FILE_PATH)
        workbook.close()
        print(f"已创建示例数据文件: {DATA_FILE_PATH}")
    
    except Exception as e:
        print(f"创建示例文件错误: {e}")


# ==================== 主程序入口 ====================

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("使用方法:")
        print("  python login_helper.py md5     - 计算MD5")
        print("  python login_helper.py verify  - 验证用户")
        print("  python login_helper.py update  - 更新登录记录")
        print("  python login_helper.py create  - 创建示例数据文件")
        sys.exit(1)
    
    command = sys.argv[1].lower()
    
    if command == "md5":
        sys.exit(ProcessMd5Command())
    elif command == "verify":
        sys.exit(ProcessVerifyCommand())
    elif command == "update":
        sys.exit(ProcessUpdateCommand())
    elif command == "create":
        CreateSampleDataFile()
        sys.exit(0)
    else:
        print(f"未知命令: {command}")
        sys.exit(1)
