#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
登录辅助模块
负责读取和更新DATA.xlsx中的用户数据
"""

import sys
import os
from datetime import datetime


def ReadUserData(filename='DATA.xlsx'):
    """
    传入值: filename (str) - Excel文件名
    返回值: dict - 用户数据字典 {username: {password_md5, last_login, last_ip}}
    
    功能: 读取DATA.xlsx中的用户数据
    """
    user_data = {}
    
    try:
        # 检查是否安装了openpyxl
        try:
            from openpyxl import load_workbook
        except ImportError:
            # 如果没有安装openpyxl，使用简单的文本文件模拟
            print("警告: 未安装openpyxl，使用文本文件模拟", file=sys.stderr)
            return ReadUserDataFromText('DATA.txt')
        
        if not os.path.exists(filename):
            print(f"警告: {filename} 不存在，使用默认数据", file=sys.stderr)
            return GetDefaultUserData()
        
        wb = load_workbook(filename)
        ws = wb.active
        
        # 假设第一行是表头: 用户名, 密码MD5, 最后登录时间, 最后登录IP
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:  # 确保用户名不为空
                username = str(row[0])
                password_md5 = str(row[1]) if row[1] else ""
                last_login = str(row[2]) if row[2] else ""
                last_ip = str(row[3]) if row[3] else ""
                
                user_data[username] = {
                    'password_md5': password_md5,
                    'last_login': last_login,
                    'last_ip': last_ip
                }
        
        wb.close()
        
    except Exception as e:
        print(f"读取Excel错误: {e}", file=sys.stderr)
        return GetDefaultUserData()
    
    return user_data


def ReadUserDataFromText(filename='DATA.txt'):
    """
    传入值: filename (str) - 文本文件名
    返回值: dict - 用户数据字典
    
    功能: 从文本文件读取用户数据（作为Excel的备选方案）
    """
    user_data = {}
    
    if not os.path.exists(filename):
        return GetDefaultUserData()
    
    try:
        with open(filename, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#'):
                    parts = line.split('|')
                    if len(parts) >= 2:
                        username = parts[0].strip()
                        password_md5 = parts[1].strip()
                        last_login = parts[2].strip() if len(parts) > 2 else ""
                        last_ip = parts[3].strip() if len(parts) > 3 else ""
                        
                        user_data[username] = {
                            'password_md5': password_md5,
                            'last_login': last_login,
                            'last_ip': last_ip
                        }
    except Exception as e:
        print(f"读取文本文件错误: {e}", file=sys.stderr)
        return GetDefaultUserData()
    
    return user_data


def GetDefaultUserData():
    """
    传入值: 无
    返回值: dict - 默认测试用户数据
    
    功能: 返回默认的测试用户数据
    """
    # 导入MD5函数计算测试密码
    from algorithm import CalculateMD5
    
    return {
        'admin': {
            'password_md5': CalculateMD5('admin123'),
            'last_login': '',
            'last_ip': ''
        },
        'user123': {
            'password_md5': CalculateMD5('pass1234'),
            'last_login': '',
            'last_ip': ''
        },
        'test': {
            'password_md5': CalculateMD5('test1234'),
            'last_login': '',
            'last_ip': ''
        }
    }


def UpdateUserLoginInfo(filename, username, login_time, ip_address):
    """
    传入值: filename (str) - Excel文件名, username (str) - 用户名, 
           login_time (str) - 登录时间, ip_address (str) - IP地址
    返回值: bool - 是否更新成功
    
    功能: 更新用户的最后登录时间和IP地址到DATA.xlsx
    """
    try:
        try:
            from openpyxl import load_workbook, Workbook
        except ImportError:
            # 使用文本文件
            return UpdateUserLoginInfoToText('DATA.txt', username, login_time, ip_address)
        
        # 如果文件不存在，创建新文件
        if not os.path.exists(filename):
            wb = Workbook()
            ws = wb.active
            ws.title = "Users"
            # 添加表头
            ws.append(['用户名', '密码MD5', '最后登录时间', '最后登录IP'])
            # 添加默认用户
            from algorithm import CalculateMD5
            ws.append(['admin', CalculateMD5('admin123'), '', ''])
            ws.append(['user123', CalculateMD5('pass1234'), '', ''])
            ws.append(['test', CalculateMD5('test1234'), '', ''])
            wb.save(filename)
        
        wb = load_workbook(filename)
        ws = wb.active
        
        # 查找用户并更新
        updated = False
        for row in ws.iter_rows(min_row=2):
            if row[0].value == username:
                row[2].value = login_time
                row[3].value = ip_address
                updated = True
                break
        
        if updated:
            wb.save(filename)
            wb.close()
            return True
        else:
            wb.close()
            return False
            
    except Exception as e:
        print(f"更新Excel错误: {e}", file=sys.stderr)
        return False


def UpdateUserLoginInfoToText(filename, username, login_time, ip_address):
    """
    传入值: filename (str) - 文本文件名, username (str) - 用户名,
           login_time (str) - 登录时间, ip_address (str) - IP地址
    返回值: bool - 是否更新成功
    
    功能: 更新文本文件中的用户登录信息
    """
    try:
        user_data = ReadUserDataFromText(filename)
        
        if username in user_data:
            user_data[username]['last_login'] = login_time
            user_data[username]['last_ip'] = ip_address
            
            # 写回文件
            with open(filename, 'w', encoding='utf-8') as f:
                f.write("# 用户名 | 密码MD5 | 最后登录时间 | 最后登录IP\n")
                for uname, info in user_data.items():
                    f.write(f"{uname}|{info['password_md5']}|{info['last_login']}|{info['last_ip']}\n")
            
            return True
        else:
            return False
            
    except Exception as e:
        print(f"更新文本文件错误: {e}", file=sys.stderr)
        return False


def VerifyUserLogin(username, password_md5, login_time, ip_address):
    """
    传入值: username (str) - 用户名, password_md5 (str) - 密码MD5值,
           login_time (str) - 登录时间, ip_address (str) - IP地址
    返回值: NULL (通过标准输出返回结果)
    
    功能: 验证用户登录并更新登录信息
    """
    # 读取用户数据
    user_data = ReadUserData('DATA.xlsx')
    
    # 验证用户名是否存在
    if username not in user_data:
        print("FAILED:用户名不存在")
        return
    
    # 验证密码MD5
    if user_data[username]['password_md5'] != password_md5:
        print("FAILED:密码错误")
        return
    
    # 验证成功，更新登录信息
    if UpdateUserLoginInfo('DATA.xlsx', username, login_time, ip_address):
        print("SUCCESS:登录成功")
    else:
        print("SUCCESS:登录成功（但更新登录信息失败）")


def InitializeDataFile():
    """
    传入值: 无
    返回值: NULL
    
    功能: 初始化DATA.xlsx和DATA.txt文件
    """
    try:
        from openpyxl import Workbook
        from algorithm import CalculateMD5
        
        # 创建Excel文件
        wb = Workbook()
        ws = wb.active
        ws.title = "Users"
        
        # 添加表头
        ws.append(['用户名', '密码MD5', '最后登录时间', '最后登录IP'])
        
        # 添加测试用户
        ws.append(['admin', CalculateMD5('admin123'), '', ''])
        ws.append(['user123', CalculateMD5('pass1234'), '', ''])
        ws.append(['test', CalculateMD5('test1234'), '', ''])
        
        wb.save('DATA.xlsx')
        print("DATA.xlsx 创建成功")
        
    except ImportError:
        print("未安装openpyxl，创建文本文件版本")
    
    # 创建文本文件版本
    from algorithm import CalculateMD5
    
    with open('DATA.txt', 'w', encoding='utf-8') as f:
        f.write("# 用户名 | 密码MD5 | 最后登录时间 | 最后登录IP\n")
        f.write(f"admin|{CalculateMD5('admin123')}||\n")
        f.write(f"user123|{CalculateMD5('pass1234')}||\n")
        f.write(f"test|{CalculateMD5('test1234')}||\n")
    
    print("DATA.txt 创建成功")


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("用法: python login_helper.py <command> [args...]")
        print("命令:")
        print("  verify <username> <password_md5> <login_time> <ip_address> - 验证登录")
        print("  init - 初始化数据文件")
        sys.exit(1)
    
    command = sys.argv[1]
    
    if command == 'verify':
        if len(sys.argv) != 6:
            print("FAILED:参数错误")
            sys.exit(1)
        
        username = sys.argv[2]
        password_md5 = sys.argv[3]
        login_time = sys.argv[4]
        ip_address = sys.argv[5]
        
        VerifyUserLogin(username, password_md5, login_time, ip_address)
    
    elif command == 'init':
        InitializeDataFile()
    
    else:
        print(f"未知命令: {command}")
        sys.exit(1)
