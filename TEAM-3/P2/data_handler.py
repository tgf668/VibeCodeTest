"""
数据处理辅助模块 - 处理Excel数据的读写操作
供C语言模块调用
"""

import sys
import json
from datetime import datetime
import os
import re

# 尝试导入openpyxl库，如果不存在则使用备用方案
try:
    from openpyxl import load_workbook, Workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("警告: openpyxl未安装，将使用JSON文件作为数据存储", file=sys.stderr)


# 常量定义
DATA_FILE_XLSX = 'DATA.xlsx'
DATA_FILE_JSON = 'DATA.json'  # 备用JSON存储


def ValidateUsername(username):
    """
    验证用户名的合法性 - 修复CWE-20: 输入验证
    传入值：username (str) - 用户名
    返回值：bool - 合法返回True
    """
    if not username or not isinstance(username, str):
        return False
    
    # 限制用户名长度和字符
    if len(username) > 100:
        return False
    
    # 只允许字母、数字、下划线
    if not re.match(r'^[a-zA-Z0-9_]+$', username):
        return False
    
    return True


def ValidateIpAddress(ip_address):
    """
    验证IP地址的合法性 - 修复CWE-20: 输入验证
    传入值：ip_address (str) - IP地址
    返回值：bool - 合法返回True
    """
    if not ip_address or not isinstance(ip_address, str):
        return False
    
    # 验证IPv4地址格式
    ipv4_pattern = r'^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$'
    if not re.match(ipv4_pattern, ip_address):
        return False
    
    # 验证每个八位组在0-255范围内
    parts = ip_address.split('.')
    for part in parts:
        if int(part) > 255:
            return False
    
    return True


def ReadUserData(username):
    """
    从数据文件中读取用户数据
    传入值：username (str) - 用户名
    返回值：dict - 用户数据（包含password_md5等），未找到返回None
    """
    # 修复CWE-20: 输入验证
    if not ValidateUsername(username):
        print(f"错误: 无效的用户名: {username}", file=sys.stderr)
        return None
    
    if EXCEL_AVAILABLE and os.path.exists(DATA_FILE_XLSX):
        return ReadUserDataFromExcel(username)
    else:
        return ReadUserDataFromJson(username)


def ReadUserDataFromExcel(username):
    """
    从Excel文件读取用户数据
    传入值：username (str) - 用户名
    返回值：dict - 用户数据，未找到返回None
    """
    try:
        # 检查文件是否存在
        if not os.path.exists(DATA_FILE_XLSX):
            # 创建新的Excel文件
            CreateDefaultExcelFile()
        
        # 加载工作簿
        workbook = load_workbook(DATA_FILE_XLSX)
        sheet = workbook.active
        
        # 遍历所有行查找用户（从第2行开始，第1行是表头）
        for row in sheet.iter_rows(min_row=2, values_only=False):
            if row[0].value == username:
                # 找到用户，返回用户数据
                return {
                    'username': row[0].value,
                    'password_md5': row[1].value,
                    'last_login_time': str(row[2].value) if row[2].value else '',
                    'last_login_ip': row[3].value if row[3].value else '',
                    'row_number': row[0].row  # 保存行号，用于后续更新
                }
        
        workbook.close()
        return None
        
    except Exception as e:
        print(f"Excel读取错误: {str(e)}", file=sys.stderr)
        return None


def ReadUserDataFromJson(username):
    """
    从JSON文件读取用户数据（备用方案）
    传入值：username (str) - 用户名
    返回值：dict - 用户数据，未找到返回None
    """
    try:
        if not os.path.exists(DATA_FILE_JSON):
            CreateDefaultJsonFile()
        
        with open(DATA_FILE_JSON, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        for user in data.get('users', []):
            if user['username'] == username:
                return user
        
        return None
        
    except Exception as e:
        print(f"JSON读取错误: {str(e)}", file=sys.stderr)
        return None


def UpdateUserLoginInfo(username, login_ip):
    """
    更新用户的登录信息
    传入值：username (str) - 用户名
            login_ip (str) - 登录IP地址
    返回值：bool - 成功返回True，失败返回False
    """
    # 修复CWE-20: 输入验证
    if not ValidateUsername(username):
        print(f"错误: 无效的用户名: {username}", file=sys.stderr)
        return False
    
    if not ValidateIpAddress(login_ip):
        print(f"错误: 无效的IP地址: {login_ip}", file=sys.stderr)
        return False
    
    if EXCEL_AVAILABLE and os.path.exists(DATA_FILE_XLSX):
        return UpdateUserLoginInfoInExcel(username, login_ip)
    else:
        return UpdateUserLoginInfoInJson(username, login_ip)


def UpdateUserLoginInfoInExcel(username, login_ip):
    """
    在Excel文件中更新用户登录信息
    传入值：username (str) - 用户名
            login_ip (str) - 登录IP地址
    返回值：bool - 成功返回True，失败返回False
    """
    try:
        workbook = load_workbook(DATA_FILE_XLSX)
        sheet = workbook.active
        
        # 获取当前时间
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # 查找用户并更新登录信息
        for row in sheet.iter_rows(min_row=2, values_only=False):
            if row[0].value == username:
                row[2].value = current_time  # 更新登录时间
                row[3].value = login_ip       # 更新登录IP
                break
        
        # 保存工作簿
        workbook.save(DATA_FILE_XLSX)
        workbook.close()
        return True
        
    except Exception as e:
        print(f"Excel更新错误: {str(e)}", file=sys.stderr)
        return False


def UpdateUserLoginInfoInJson(username, login_ip):
    """
    在JSON文件中更新用户登录信息（备用方案）
    传入值：username (str) - 用户名
            login_ip (str) - 登录IP地址
    返回值：bool - 成功返回True，失败返回False
    """
    try:
        with open(DATA_FILE_JSON, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        # 获取当前时间
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # 查找用户并更新
        for user in data.get('users', []):
            if user['username'] == username:
                user['last_login_time'] = current_time
                user['last_login_ip'] = login_ip
                break
        
        # 保存回文件
        with open(DATA_FILE_JSON, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        
        return True
        
    except Exception as e:
        print(f"JSON更新错误: {str(e)}", file=sys.stderr)
        return False


def CreateDefaultExcelFile():
    """
    创建默认的Excel数据文件
    传入值：None
    返回值：None
    """
    try:
        workbook = Workbook()
        sheet = workbook.active
        
        # 设置表头
        sheet['A1'] = '用户名'
        sheet['B1'] = '密码MD5'
        sheet['C1'] = '最后登录时间'
        sheet['D1'] = '最后登录IP'
        
        # 添加默认测试用户（密码123456的MD5值）
        sheet.append(['admin', 'e10adc3949ba59abbe56e057f20f883e', '', ''])
        sheet.append(['test', '098f6bcd4621d373cade4e832627b4f6', '', ''])  # test的MD5
        
        workbook.save(DATA_FILE_XLSX)
        workbook.close()
        print(f"已创建默认Excel文件: {DATA_FILE_XLSX}")
        
    except Exception as e:
        print(f"创建Excel文件错误: {str(e)}", file=sys.stderr)


def CreateDefaultJsonFile():
    """
    创建默认的JSON数据文件（备用方案）
    传入值：None
    返回值：None
    """
    try:
        default_data = {
            'users': [
                {
                    'username': 'admin',
                    'password_md5': 'e10adc3949ba59abbe56e057f20f883e',  # 123456
                    'last_login_time': '',
                    'last_login_ip': ''
                },
                {
                    'username': 'test',
                    'password_md5': '098f6bcd4621d373cade4e832627b4f6',  # test
                    'last_login_time': '',
                    'last_login_ip': ''
                }
            ]
        }
        
        with open(DATA_FILE_JSON, 'w', encoding='utf-8') as f:
            json.dump(default_data, f, ensure_ascii=False, indent=2)
        
        print(f"已创建默认JSON文件: {DATA_FILE_JSON}")
        
    except Exception as e:
        print(f"创建JSON文件错误: {str(e)}", file=sys.stderr)


if __name__ == '__main__':
    # 命令行接口
    if len(sys.argv) < 2:
        print("用法: python data_handler.py <command> [args]")
        print("命令:")
        print("  read <username>              - 读取用户数据")
        print("  update <username> <ip>       - 更新登录信息")
        print("  init                         - 初始化数据文件")
        sys.exit(1)
    
    command = sys.argv[1]
    
    if command == 'read' and len(sys.argv) >= 3:
        username = sys.argv[2]
        user_data = ReadUserData(username)
        if user_data:
            print(json.dumps(user_data, ensure_ascii=False))
        else:
            print(json.dumps({'error': 'User not found'}))
            sys.exit(1)
    
    elif command == 'update' and len(sys.argv) >= 4:
        username = sys.argv[2]
        login_ip = sys.argv[3]
        success = UpdateUserLoginInfo(username, login_ip)
        if success:
            print(json.dumps({'status': 'OK'}))
        else:
            print(json.dumps({'status': 'ERROR'}))
            sys.exit(1)
    
    elif command == 'init':
        if EXCEL_AVAILABLE:
            CreateDefaultExcelFile()
        else:
            CreateDefaultJsonFile()
    
    else:
        print("无效命令")
        sys.exit(1)
